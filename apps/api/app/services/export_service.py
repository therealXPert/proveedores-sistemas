"""
Exportaciones (sección 15 del diseño): CSV, Excel y PDF de resultados
filtrados. Las exportaciones respetan los mismos filtros que ya aplico
el usuario en pantalla (los datos ya vienen filtrados desde el endpoint).
"""
import io
import csv
from datetime import datetime

import openpyxl
from openpyxl.styles import Font
from fpdf import FPDF


def _rows_to_csv(rows: list[dict], columnas: list[tuple[str, str]]) -> bytes:
    """columnas: lista de (clave, encabezado)."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow([enc for _, enc in columnas])
    for row in rows:
        writer.writerow([row.get(clave, "") for clave, _ in columnas])
    return buffer.getvalue().encode("utf-8-sig")  # BOM para que Excel abra bien los acentos


def _rows_to_xlsx(rows: list[dict], columnas: list[tuple[str, str]], titulo: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # limite de Excel para nombres de hoja

    for col_idx, (_, encabezado) in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=encabezado)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (clave, _) in enumerate(columnas, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(clave, ""))

    for col_idx, (clave, encabezado) in enumerate(columnas, start=1):
        ancho = max(len(encabezado), *(len(str(r.get(clave, ""))) for r in rows)) if rows else len(encabezado)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(ancho + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


INVOICE_COLUMNAS = [
    ("fecha_emision", "Fecha"),
    ("tipo_documento", "Tipo de Documento"),
    ("numero_factura", "Número de Factura"),
    ("provider_nombre", "Proveedor"),
    ("area_nombre", "Área"),
    ("category_nombre", "Categoría"),
    ("importe_total", "Importe Total"),
    ("moneda", "Moneda"),
    ("descripcion", "Descripción"),
    ("usuario_aprobador_nombre", "Aprobado por"),
]


def invoices_to_csv(invoices: list[dict]) -> bytes:
    return _rows_to_csv(invoices, INVOICE_COLUMNAS)


def invoices_to_xlsx(invoices: list[dict]) -> bytes:
    return _rows_to_xlsx(invoices, INVOICE_COLUMNAS, "Facturas")


BUDGET_RESUMEN_COLUMNAS = [
    ("provider_nombre", "Proveedor"),
    ("presupuesto_mensual", "Presupuesto Mensual"),
    ("gasto_real_promedio_mensual", "Gasto Real Promedio Mensual"),
    ("gasto_real_total", "Gasto Real Total"),
    ("desvio_mensual", "Desvío Mensual"),
]


def budget_resumen_to_csv(resumen: list[dict]) -> bytes:
    return _rows_to_csv(resumen, BUDGET_RESUMEN_COLUMNAS)


def budget_resumen_to_xlsx(resumen: list[dict]) -> bytes:
    return _rows_to_xlsx(resumen, BUDGET_RESUMEN_COLUMNAS, "Presupuesto vs Real")


def _fmt_monto(v: float) -> str:
    return f"{v:,.0f}".replace(",", ".")


class InformeEjecutivoPDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 10, "Control de Gasto - Sistemas", ln=True)
        self.set_font("Helvetica", "", 10)
        self.set_text_color(100, 100, 100)
        self.cell(0, 6, "Informe ejecutivo - Autocity", ln=True)
        self.set_text_color(0, 0, 0)
        self.ln(4)


def dashboard_to_pdf(kpis: dict, top_proveedores: list[dict]) -> bytes:
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
             "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    pdf = InformeEjecutivoPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, f"Período: {meses[kpis['mes'] - 1]} {kpis['anio']}", ln=True)
    pdf.cell(0, 6, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True)
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Indicadores principales", ln=True)
    pdf.set_font("Helvetica", "", 10)

    filas = [
        ("Gasto del mes", f"$ {_fmt_monto(kpis['gasto_total_mes'])}"),
        ("Gasto acumulado del año", f"$ {_fmt_monto(kpis['gasto_acumulado_anio'])}"),
        ("Presupuesto mensual", f"$ {_fmt_monto(kpis['presupuesto_mensual'])}"),
        ("Presupuesto anual", f"$ {_fmt_monto(kpis['presupuesto_anual'])}"),
        ("% presupuesto consumido", f"{kpis['porcentaje_consumido_mes']:.1f}%" if kpis["porcentaje_consumido_mes"] is not None else "-"),
        ("Desvío contra presupuesto (mes)", f"$ {_fmt_monto(kpis['desvio_contra_presupuesto_mes'])}"),
        ("Proyección de cierre del año", f"$ {_fmt_monto(kpis['proyeccion_cierre_anio'])}" if kpis["proyeccion_cierre_anio"] is not None else "-"),
        ("Variación vs. mes anterior", f"{kpis['variacion_mes_anterior_pct']:.1f}%" if kpis["variacion_mes_anterior_pct"] is not None else "-"),
        ("Cantidad de facturas del mes", str(kpis["cantidad_facturas_mes"])),
        ("Cantidad de proveedores del mes", str(kpis["cantidad_proveedores_mes"])),
    ]
    for etiqueta, valor in filas:
        pdf.cell(90, 7, etiqueta, border=0)
        pdf.cell(0, 7, valor, border=0, ln=True)

    pdf.ln(6)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Principales proveedores del período", ln=True)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(230, 230, 230)
    pdf.cell(90, 7, "Proveedor", border=1, fill=True)
    pdf.cell(45, 7, "Monto", border=1, fill=True, align="R")
    pdf.cell(30, 7, "% Particip.", border=1, fill=True, align="R", ln=True)
    pdf.set_font("Helvetica", "", 9)
    for p in top_proveedores[:10]:
        pdf.cell(90, 6, p["nombre"][:45], border=1)
        pdf.cell(45, 6, _fmt_monto(p["monto"]), border=1, align="R")
        pdf.cell(30, 6, f"{p['participacion_pct']:.1f}%", border=1, align="R", ln=True)

    return bytes(pdf.output())
