"""Tests de exportacion (CSV/Excel/PDF) y estructura basica de auditoria."""
import csv
import io

import openpyxl

from app.services.export_service import invoices_to_csv, invoices_to_xlsx, dashboard_to_pdf


SAMPLE_INVOICES = [
    {
        "fecha_emision": "2026-01-15T00:00:00",
        "tipo_documento": "Factura",
        "numero_factura": "0001-00000123",
        "provider_nombre": "Proveedor Test",
        "area_nombre": "Sistemas Mesa de Servicio",
        "category_nombre": "Licencias",
        "importe_total": 150000.50,
        "moneda": "Pesos",
        "descripcion": "Licencias de software",
        "usuario_aprobador_nombre": "Administrador",
    }
]

SAMPLE_KPIS = {
    "fecha_desde": "2026-07-01",
    "fecha_hasta": "2026-07-31",
    "dias": 31,
    "gasto_total_periodo": 1000000.0,
    "presupuesto_periodo": 900000.0,
    "porcentaje_consumido": 111.1,
    "desvio_contra_presupuesto": 100000.0,
    "variacion_vs_periodo_anterior_pct": 5.2,
    "cantidad_facturas": 42,
    "cantidad_proveedores": 15,
    "importaciones_pendientes": 0,
    "registros_con_error": 0,
}

SAMPLE_TOP_PROVEEDORES = [
    {"provider_id": 1, "nombre": "Proveedor Test", "monto": 500000.0, "participacion_pct": 50.0, "acumulado_pct": 50.0},
]


def test_invoices_to_csv_contiene_encabezados_y_datos():
    contenido = invoices_to_csv(SAMPLE_INVOICES)
    texto = contenido.decode("utf-8-sig")
    reader = list(csv.reader(io.StringIO(texto), delimiter=";"))

    assert reader[0] == [
        "Fecha", "Tipo de Documento", "Número de Factura", "Proveedor", "Área",
        "Categoría", "Importe Total", "Moneda", "Descripción", "Aprobado por",
    ]
    assert reader[1][3] == "Proveedor Test"
    assert reader[1][6] == "150000.5"


def test_invoices_to_xlsx_es_un_excel_valido_con_los_datos():
    contenido = invoices_to_xlsx(SAMPLE_INVOICES)
    wb = openpyxl.load_workbook(io.BytesIO(contenido))
    ws = wb.active

    assert ws["A1"].value == "Fecha"
    assert ws["D2"].value == "Proveedor Test"


def test_dashboard_to_pdf_genera_bytes_de_pdf_valido():
    contenido = dashboard_to_pdf(SAMPLE_KPIS, SAMPLE_TOP_PROVEEDORES)
    assert contenido[:4] == b"%PDF"
    assert len(contenido) > 500


def test_dashboard_to_pdf_sin_top_proveedores_no_falla():
    contenido = dashboard_to_pdf(SAMPLE_KPIS, [])
    assert contenido[:4] == b"%PDF"
